import sys
import os
import json
import PyPDF2
import docx
import re
import openpyxl

# Khai báo đường dẫn để Python hiểu cấu trúc thư mục
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

# Import đầy đủ các Parser và Grader cho 3 định dạng
from Parser.docx_parser.core import parse_docx
from Grade.word_grade import grade_submission as grade_word

from Parser.excel_parser.core import parse_xlsx
from Grade.excel_grade import grade_submission as grade_excel

from Parser.pptx_parser.core import parse_pptx
from Grade.pptx_grade import grade_submission as grade_pptx

from result_to_pdf import export_result_to_pdf_from_dict


def extract_file_text(file_path):
    text = ""
    try:
        if file_path.lower().endswith('.pdf'):
            import PyPDF2
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    if page.extract_text():
                        text += page.extract_text() + "\n"
        elif file_path.lower().endswith('.docx'):
            import docx
            doc = docx.Document(file_path)
            for para in doc.paragraphs:
                text += para.text + "\n"
        else:
            return {"error": "Định dạng file đề không được hỗ trợ để trích xuất chữ."}
            
        return {"text": text.strip()}
    except Exception as e:
        return {"error": str(e)}

def load_students_info(excel_path):
    """Đọc file Excel danh sách sinh viên và trả về Dict mapping theo MSSV"""
    student_dict = {}
    if not os.path.exists(excel_path):
        return student_dict
        
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active
        # Lấy header dòng 1 và tìm vị trí cột
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in sheet[1]]
        
        idx_mssv, idx_name, idx_class = -1, -1, -1
        for i, h in enumerate(headers):
            if 'mssv' in h or 'mã' in h: idx_mssv = i
            elif 'tên' in h or 'name' in h: idx_name = i
            elif 'lớp' in h or 'class' in h: idx_class = i
            
        # Đọc dữ liệu từ dòng 2
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if idx_mssv != -1 and row[idx_mssv]:
                mssv = str(row[idx_mssv]).strip()
                student_dict[mssv] = {
                    "name": str(row[idx_name]).strip() if idx_name != -1 and row[idx_name] else "N/A",
                    "class": str(row[idx_class]).strip() if idx_class != -1 and row[idx_class] else "N/A"
                }
    except Exception as e:
        print(f"Lỗi đọc file sinh viên: {e}")
        
    return student_dict

def main():
    # SỬA LỖI Ở ĐÂY: Chỉ check < 2 để cho phép lệnh grade_local đi qua
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Thiếu tham số dòng lệnh."}))
        sys.exit(0)

    command = sys.argv[1].lower()

    try:
        if command == "grade_local":
            submit_dir = "/shared_workspace/submit"
            rubric_dir = "/shared_workspace/rubric"
            result_dir = "/shared_workspace/result"
            student_file = "/shared_workspace/students/students.xlsx"

            os.makedirs(submit_dir, exist_ok=True)
            os.makedirs(rubric_dir, exist_ok=True)
            os.makedirs(result_dir, exist_ok=True)

            # 1. Load danh sách sinh viên
            students_db = load_students_info(student_file)

            # 2. Load và phân loại Rubrics
            rubrics_map = {
                '.docx': [],
                '.xlsx': [],
                '.pptx': []
            }

            for r_file in os.listdir(rubric_dir):
                if r_file.endswith('.json'):
                    r_path = os.path.join(rubric_dir, r_file)

                    with open(r_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)

                    name_lower = r_file.lower()

                    if 'word' in name_lower or 'docx' in name_lower:
                        rubrics_map['.docx'].append(data)

                    elif 'excel' in name_lower or 'xlsx' in name_lower:
                        rubrics_map['.xlsx'].append(data)

                    elif 'powerpoint' in name_lower or 'pptx' in name_lower or 'power' in name_lower:
                        rubrics_map['.pptx'].append(data)

            if not any(rubrics_map.values()):
                print(json.dumps({"error": "Không tìm thấy rubric."}))
                sys.exit(0)

            summary_data = []

            # 3. Quét đệ quy (os.walk) để giữ nguyên cấu trúc thư mục
            for root, dirs, files in os.walk(submit_dir):
                for s_file in files:

                    if s_file.startswith('~') or s_file.lower() == 'thumbs.db':
                        continue

                    ext = os.path.splitext(s_file)[1].lower()

                    if ext not in ['.docx', '.xlsx', '.pptx']:
                        continue

                    s_path = os.path.join(root, s_file)

                    rel_path = os.path.relpath(root, submit_dir)

                    current_out_dir = os.path.join(result_dir, rel_path)
                    os.makedirs(current_out_dir, exist_ok=True)

                    pdf_result_path = os.path.join(current_out_dir, f"{s_file}.pdf")

                    if rel_path == ".":
                        display_path = f"{s_file}.pdf"
                    else:
                        display_path = f"{rel_path}/{s_file}.pdf".replace("\\", "/")

                    mssv = "Unknown"

                    mssv_match = re.search(r'\d{8}', s_file)

                    if not mssv_match:
                        mssv_match = re.search(r'\d{8}', os.path.basename(root))

                    if mssv_match:
                        mssv = mssv_match.group(0)

                    s_info = students_db.get(mssv, {
                        "name": "N/A",
                        "class": "N/A"
                    })

                    info_string = f"{mssv}-{s_info['name']}-{s_info['class']}"

                    file_summary = {
                        "class": s_info['class'],
                        "mssv": mssv,
                        "name": s_info['name'],
                        "ext": ext,
                        "file_name": s_file,
                        "status": "processing",
                        "result_path": display_path
                    }

                    try:

                        if ext not in rubrics_map or not rubrics_map[ext]:
                            raise ValueError(f"Không có rubric cho {ext}")

                        rubric_list = rubrics_map[ext]

                        # Parse AST chỉ 1 lần
                        if ext == '.docx':
                            ast = parse_docx(s_path, clean=True)

                        elif ext == '.xlsx':
                            ast = parse_xlsx(s_path, clean=True)

                        elif ext == '.pptx':
                            ast = parse_pptx(s_path, clean=True)

                        best_result = None
                        best_score = -1

                        # Chấm tất cả rubric và lấy kết quả tốt nhất
                        for rubric in rubric_list:

                            if ext == '.docx':
                                result = grade_word(rubric, ast)

                            elif ext == '.xlsx':
                                result = grade_excel(rubric, ast)

                            elif ext == '.pptx':
                                result = grade_pptx(rubric, ast)

                            score = result.get("final_score", 0)

                            if score > best_score:
                                best_score = score
                                best_result = result

                        if best_result is None:
                            raise ValueError("Không rubric nào chấm được file")

                        grade_result = best_result

                        # Xuất PDF
                        export_result_to_pdf_from_dict(
                            info_string,
                            grade_result,
                            pdf_result_path
                        )

                        # Trích xuất Metadata
                        props = grade_result.get("properties", {})
                        meta = props.get("metadata", {})
                        core = props.get("core_properties", {})

                        created_by = (
                            meta.get("creator")
                            or core.get("creator")
                            or core.get("dc:creator")
                            or ""
                        )

                        updated_by = (
                            meta.get("lastModifiedBy")
                            or core.get("lastModifiedBy")
                            or core.get("cp:lastModifiedBy")
                            or ""
                        )

                        file_summary.update({
                            "status": "SUCCESS",
                            "final_score": grade_result.get("final_score", 0),
                            "created_by": created_by,
                            "updated_by": updated_by
                        })

                    except Exception as e:

                        file_summary.update({
                            "status": "FAILED",
                            "error_msg": str(e),
                            "final_score": 0,
                            "created_by": "",
                            "updated_by": ""
                        })

                    summary_data.append(file_summary)

            # 4. Tạo file Excel tổng hợp cho Giảng viên
            summary_wb = openpyxl.Workbook()

            summary_wb.remove(summary_wb.active)

            headers = [
                "Lớp",
                "MSSV",
                "Họ Tên",
                "Điểm",
                "Path Kết Quả (PDF)",
                "Tạo Bởi (Meta)",
                "Sửa Lần Cuối (Meta)",
                "Trạng thái lỗi"
            ]

            sheets_map = {
                'Word': (
                    '.docx',
                    summary_wb.create_sheet(title="Word")
                ),
                'Excel': (
                    '.xlsx',
                    summary_wb.create_sheet(title="Excel")
                ),
                'PowerPoint': (
                    '.pptx',
                    summary_wb.create_sheet(title="PowerPoint")
                )
            }

            for _, (_, ws) in sheets_map.items():
                ws.append(headers)

            for item in summary_data:

                for sheet_name, (ext, ws) in sheets_map.items():

                    if item['ext'] == ext:

                        ws.append([
                            item.get("class"),
                            item.get("mssv"),
                            item.get("name"),
                            item.get("final_score"),
                            item.get("result_path"),
                            item.get("created_by"),
                            item.get("updated_by"),
                            item.get(
                                "error_msg",
                                "OK" if item.get("status") == "SUCCESS"
                                else item.get("error_msg")
                            )
                        ])

            summary_wb.save(
                os.path.join(result_dir, "TongHopDiem.xlsx")
            )

            # 5. Trả JSON tổng hợp ra cho n8n
            print(json.dumps(summary_data, ensure_ascii=False))

        # ==========================================
        # CÁC LỆNH CŨ (PARSE, EXTRACT_TEXT, GRADE LẺ)
        # ==========================================
        else:
            # SỬA LỖI Ở ĐÂY: Dời việc lấy file_path vào trong block này
            if len(sys.argv) < 3:
                print(json.dumps({"error": "Thiếu tham số dòng lệnh. Cú pháp: python n8n_bridge.py [parse/grade] [file_path]"}))
                sys.exit(0)
            
            file_path = sys.argv[2]

            if not os.path.exists(file_path):
                print(json.dumps({"error": f"Không tìm thấy file tại: {file_path}"}))
                sys.exit(0)

            if command == "parse":
                ast = None
                if file_path.lower().endswith('.docx'):
                    ast = parse_docx(file_path, clean=True)
                elif file_path.lower().endswith('.xlsx'):
                    ast = parse_xlsx(file_path, clean=True)
                elif file_path.lower().endswith('.pptx'):
                    ast = parse_pptx(file_path, clean=True)
                else:
                    print(json.dumps({"error": "Định dạng file không được hỗ trợ để parse."}))
                    sys.exit(0)
                
                print(json.dumps(ast, ensure_ascii=False))

            elif command == "extract_text":
                result = extract_file_text(file_path)
                print(json.dumps(result, ensure_ascii=False))

            elif command == "grade":
                if len(sys.argv) < 4:
                    print(json.dumps({"error": "Thiếu file rubric. Cú pháp: python n8n_bridge.py grade [file_path] [rubric_path]"}))
                    sys.exit(0)
                    
                rubric_path = sys.argv[3]
                if not os.path.exists(rubric_path):
                    print(json.dumps({"error": f"Không tìm thấy file rubric tại: {rubric_path}"}))
                    sys.exit(0)

                with open(rubric_path, 'r', encoding='utf-8') as f:
                    rubric = json.load(f)
                
                result = None
                if file_path.lower().endswith('.docx'):
                    ast = parse_docx(file_path, clean=True)
                    result = grade_word(rubric, ast)
                elif file_path.lower().endswith('.xlsx'):
                    ast = parse_xlsx(file_path, clean=True)
                    result = grade_excel(rubric, ast)
                elif file_path.lower().endswith('.pptx'):
                    ast = parse_pptx(file_path, clean=True)
                    result = grade_pptx(rubric, ast)
                else:
                    print(json.dumps({"error": "Định dạng file không được hỗ trợ để chấm điểm."}))
                    sys.exit(0)
                
                props = result.get("properties", {})
                metadata = props.get("metadata", {})
                core_props = props.get("core_properties", {})

                created_by = metadata.get("creator") or core_props.get("creator") or core_props.get("dc:creator") or ""
                updated_by = metadata.get("lastModifiedBy") or core_props.get("lastModifiedBy") or core_props.get("cp:lastModifiedBy") or ""
                
                base_filename = os.path.splitext(os.path.basename(file_path))[0]
                grade_dir = "/shared_workspace/grade"
                os.makedirs(grade_dir, exist_ok=True)
                result_path = os.path.join(grade_dir, f"result_{base_filename}.json")
                with open(result_path, 'w', encoding='utf-8') as f:
                    json.dump(result, f, ensure_ascii=False)
                summary = {
                    "file_path": result_path,
                    "final_score": result.get("final_score", 0),
                    "metadata_clean": {
                        "created_by": created_by,
                        "updated_by": updated_by
                    }
                }
                print(json.dumps(summary, ensure_ascii=False))
                
            else:
                print(json.dumps({"error": f"Lệnh không hợp lệ: {command}"}))
                sys.exit(0)

    except Exception as e:
        # Bắt toàn bộ lỗi runtime và trả về chuẩn JSON cho n8n
        print(json.dumps({"error": str(e)}, ensure_ascii=False))
        sys.exit(0)

if __name__ == "__main__":
    main()